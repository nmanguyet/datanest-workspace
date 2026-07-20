"""tracking.py — Helper for v2 tracking workbook (2-sheet model).

Sheets (the only sources of truth):
  - client_model_tracking    : 9 cols
        telco, client, client_code, client_model_code,
        target_score_table, hdfs_real_model_code,
        date_start, date_end, client_product
  - model_catalog  : 10 cols
        telco, client_model_code, model_family,
        target_score_table, hdfs_model_code,
        model_description, label_source, label_definition,
        documentation_link, notes

The optional `tracking` sheet, if present, is a DERIVED view rebuilt by
`refresh()` — never read by the lookup methods.

Usage:
    from tracking import Tracker
    t = Tracker("client_model_tracking_v2.xlsx")
    t.query(date="2025-06-01", telco="VNPT", client="HDBANK")
    t.resolve("vnpt_homecredit__cs_generic__v5.1")
    t.refresh()  # rebuild `tracking` sheet from the 2 source sheets

Convention: empty date_end = still active until today.
"""
from __future__ import annotations
import re
from datetime import date as _date, datetime
import pandas as pd


# Parsing helper only — maps a family token found in an assignment code to
# the candidate tokens used in `model_catalog.client_model_code`. This is
# code-format knowledge, not client/model data.
_FAM_EXPANSIONS = {
    "cs":         ["cs_generic", "cs_bnpl", "cs"],
    "bnpl":       ["cs_bnpl"],
    "lead":       ["leadgen"],
    "leadgen":    ["leadgen"],
    "is":         ["is_numeric", "is_range", "is"],
    "is_generic": ["is_numeric", "is_range", "is_generic"],
    "is_numeric": ["is_numeric"],
    "numeric":    ["is_numeric"],
}


def _to_date(v):
    if v is None or v == "" or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, _date) and not isinstance(v, datetime): return v
    if isinstance(v, datetime): return v.date()
    return pd.to_datetime(v).date()


def _norm(name):
    return name.strip().upper() if isinstance(name, str) else name


class Tracker:
    # Columns produced by _build_joined() (and written to the `tracking` sheet)
    OUTPUT_COLS = ["telco", "client", "client_product", "client_code",
                   "client_model_code", "target_score_table",
                   "hdfs_real_model_code", "model_family",
                   "date_start", "date_end"]

    REQUIRED_ASSIGN_COLS = ["telco", "client", "client_code", "client_model_code",
                            "target_score_table", "hdfs_real_model_code",
                            "date_start", "date_end", "client_product"]

    def __init__(self, xlsx_path: str):
        self.path = xlsx_path
        self._load()

    def _load(self):
        sheets = pd.read_excel(self.path,
                               sheet_name=["client_model_tracking", "model_catalog"],
                               dtype=str, keep_default_na=False)
        a = sheets["client_model_tracking"].fillna("").map(lambda x: str(x).strip())
        c = sheets["model_catalog"].fillna("").map(lambda x: str(x).strip())

        missing = [col for col in self.REQUIRED_ASSIGN_COLS if col not in a.columns]
        if missing:
            raise ValueError(
                f"`client_model_tracking` sheet missing required columns: {missing}. "
                f"Expected schema: {self.REQUIRED_ASSIGN_COLS}")
        if "client_model_code" not in c.columns or "model_family" not in c.columns:
            raise ValueError(
                "`model_catalog` sheet must have `client_model_code` and `model_family`.")

        for col in ("date_start", "date_end"):
            a[col] = pd.to_datetime(a[col], errors="coerce").dt.date

        self.client_model_tracking = a
        self.catalog = c
        self._catalog_codes = set(c["client_model_code"]) - {""}
        self._catalog_idx = c.set_index("client_model_code", drop=False)

    @staticmethod
    def resolve_client(name):
        """Normalize whitespace and case. No alias map — the workbook is
        canonical. If your caller uses a legacy name, fix it to match the
        `client` value present in the `client_model_tracking` sheet."""
        return _norm(name) if name else name

    def resolve(self, client_model_code):
        """Return the full catalog row for `client_model_code` (or its
        derived base) as a dict. Includes everything stored in
        `model_catalog`: telco, model_family, target_score_table,
        hdfs_model_code, model_description, label_source,
        label_definition, documentation_link, notes."""
        if not client_model_code or client_model_code == "-": return None
        if client_model_code in self._catalog_codes:
            return self._catalog_idx.loc[client_model_code].to_dict()
        base = self._derive_base(client_model_code)
        if base in self._catalog_codes:
            return self._catalog_idx.loc[base].to_dict()
        return None

    def query(self, date=None, telco=None, client=None,
              client_model_code=None, active_only=True):
        """Joined view of client_model_tracking + model_family (joined from catalog)."""
        df = self._build_joined()
        if telco:
            df = df[df["telco"].str.upper() == _norm(telco)]
        if client:
            df = df[df["client"].str.upper() == _norm(client)]
        if client_model_code:
            df = df[df["client_model_code"] == client_model_code]
        if active_only:
            d = _to_date(date) if date is not None else _date.today()
            def _ge(x): return x is None or pd.isna(x) or x >= d
            def _le(x): return x is None or pd.isna(x) or x <= d
            df = df[df["date_start"].apply(_le) & df["date_end"].apply(_ge)]
        df = df[~df["client_model_code"].isin(["", "-"])]
        return df.sort_values(["telco", "client", "date_start"]).reset_index(drop=True)

    def summary(self, date=None, telco=None):
        df = self.query(date=date, telco=telco)
        return df[["telco", "client", "client_product", "client_model_code",
                   "target_score_table", "hdfs_real_model_code",
                   "model_family"]].copy()

    def refresh(self):
        """Rebuild the `tracking` sheet from `client_model_tracking` + `model_catalog`."""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.formatting.rule import FormulaRule

        self._load()
        new_df = self._build_joined()[self.OUTPUT_COLS].copy()
        for col in new_df.columns:
            if col not in ("date_start", "date_end"):
                new_df[col] = new_df[col].fillna("")

        wb = load_workbook(self.path)
        if "tracking" in wb.sheetnames: del wb["tracking"]
        ws = wb.create_sheet("tracking")

        HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        HDR_FILL = PatternFill("solid", start_color="A4A4A4")
        ACTIVE   = PatternFill("solid", start_color="D5E8D4")
        CELL     = Font(name="Arial", size=10)
        WRAP     = Alignment(wrap_text=True, vertical="top")
        BORDER   = Border(*(Side(style="thin", color="DDDDDD"),) * 4)

        for j, col in enumerate(new_df.columns, 1):
            c = ws.cell(row=1, column=j, value=col)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = Alignment(wrap_text=True, vertical="center")
        for i, row in enumerate(new_df.itertuples(index=False), 2):
            for j, val in enumerate(row, 1):
                # pd.isna() uniformly catches None / pd.NaT / NaN
                if pd.isna(val) or val == "":
                    cv = None
                elif isinstance(val, _date) and not isinstance(val, datetime):
                    cv = val.isoformat()
                else:
                    cv = val
                c = ws.cell(row=i, column=j, value=cv)
                c.font = CELL; c.alignment = WRAP; c.border = BORDER

        widths = {"A": 8, "B": 14, "C": 30, "D": 22, "E": 50,
                  "F": 30, "G": 28, "H": 14, "I": 12, "J": 12}
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30
        last_col = get_column_letter(len(new_df.columns))
        last_row = len(new_df) + 1
        tbl = Table(displayName="tracking", ref=f"A1:{last_col}{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium7", showRowStripes=True)
        ws.add_table(tbl)
        # Active row highlighting (date_end blank → green)
        ws.conditional_formatting.add(f"A2:{last_col}{last_row}",
            FormulaRule(formula=['OR(ISBLANK($J2),$J2="",$J2="null")'], fill=ACTIVE))
        wb.save(self.path)
        return len(new_df)

    # ─── Internals ─────────────────────────────────────────────────────
    def _build_joined(self):
        df = self.client_model_tracking.copy()
        # client_product comes ONLY from client_model_tracking. Empty stays empty.
        df["client_product"] = df["client_product"].fillna("")
        # model_family joined from catalog via base-code derivation.
        df["_base"] = df["client_model_code"].apply(self._derive_base)
        fam_map = self.catalog.set_index("client_model_code")["model_family"].to_dict()
        df["model_family"] = df["_base"].apply(lambda b: fam_map.get(b, ""))
        df = df.drop(columns=["_base"])
        return df[self.OUTPUT_COLS]

    def _derive_base(self, code):
        if not code or code == "-": return ""
        if code in self._catalog_codes: return code
        if not (code.endswith("]") and ("__[" in code or "_[" in code)): return code
        m = re.match(r"^([a-z]+)_", code); telco = m.group(1) if m else ""
        sep = "__[" if "__[" in code else "_["
        suffix = code.rsplit(sep, 1)[1][:-1]
        if "_" not in suffix: return code
        head, ver = suffix.rsplit("_", 1)
        cands = []
        if "_" in head:
            cp, fs = head.rsplit("_", 1)
            for ff in _FAM_EXPANSIONS.get(fs, [fs]):
                cands.append(f"{telco}_{cp}__{ff}__{ver}")
        for ff in _FAM_EXPANSIONS.get(head, [head]):
            cands.append(f"{telco}__{ff}__{ver}")
        for c in cands:
            if c in self._catalog_codes: return c
        return cands[0] if cands else code

    def __repr__(self):
        return (f"Tracker({self.path!r}, "
                f"{len(self.client_model_tracking)} client_model_tracking, "
                f"{len(self.catalog)} models)")
